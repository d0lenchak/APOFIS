import logging
import re
import pandas as pd
from dash import Output, Input, State, html, no_update, ALL, dcc
import plotly.graph_objects as go
from dash import dash_table
import dash_bootstrap_components as dbc

from config import COLORS, get_colors, is_program_finance_account, extract_account_code
from data_loader import get_display_name, is_actual_study, get_study_phase

logger = logging.getLogger(__name__)


def register_bsb_callbacks(app, data):

    # get data
    core_df = data['core_df']
    forecast_df = data['forecast_df']
    map_df = data['map_df']
    bsb_df = data['bsb_df']
    po_df = data['po_df']
    study_daily_df = data['study_daily_df']
    quarter_cols = data['quarter_cols']
    forecast_quarter_cols = data['forecast_quarter_cols']


    # Helper functions
    def get_cdus_for_pcode(pcode):
        cdus = set()
        program_studies = core_df[core_df['P_Code'] == pcode]['Study_ID'].dropna().unique()
        for study_id in program_studies:
            match = re.match(r'\d{4}[-.](\d{3,4})$', str(study_id))
            if match:
                study_short = match.group(1)
                matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]
                if not matching_row.empty and pd.notna(matching_row.iloc[0]['CDU']):
                    cdus.add(matching_row.iloc[0]['CDU'])
        return cdus

    def expand_phase_groups(study_ids, selected_pcodes):
        """Expand phase group selections (PHASE_*) to individual study IDs."""
        if 'ALL' in study_ids:
            return study_ids

        expanded_study_ids = []
        for study_id in study_ids:
            if study_id.startswith('PHASE_'):
                # Extract phase from PHASE_PHASE_1 -> "Phase 1", PHASE_OTHER -> "Other"
                # Split only once to handle "PHASE_" prefix
                phase_part = study_id.split('PHASE_', 1)[1] if 'PHASE_' in study_id else study_id
                # Handle "OTHER" -> "Other", "PHASE_1" -> "Phase 1"
                if phase_part == 'OTHER':
                    phase_key = 'Other'
                else:
                    phase_key = phase_part.replace('_', ' ').title()
                # Get all studies matching this phase
                all_studies = core_df[core_df['P_Code'].isin(selected_pcodes)]['Study_ID'].dropna().unique()
                for study in all_studies:
                    if get_study_phase(study, study_daily_df) == phase_key:
                        expanded_study_ids.append(study)
            else:
                expanded_study_ids.append(study_id)

        return expanded_study_ids

    # ==================== BSB PAGE CALLBACKS ====================

    @app.callback(
        [Output('bsb-study-dropdown', 'options'),
         Output('bsb-study-dropdown', 'value')],
        [Input('bsb-project-dropdown', 'value'),
         Input('bsb-cdu-dropdown', 'value')],
        State('bsb-study-dropdown', 'value'))
    
    def update_bsb_studies(project_codes, cdu_selections, current_value):
        """Update BSB page study dropdown with phase buckets."""
        if not project_codes or len(project_codes) == 0:
            return [{'label': 'Select a Program first...', 'value': 'NONE'}], []

        # Get list of P_Codes from selection
        if 'ALL' in project_codes:
            selected_pcodes = core_df['P_Code'].dropna().unique().tolist()
        else:
            selected_pcodes = [pc for pc in project_codes if pc != 'ALL']

        if len(selected_pcodes) == 0:
            return [{'label': 'No programs selected', 'value': 'NONE'}], []

        # Get ALL studies for selected programs (including placeholders)
        studies = core_df[core_df['P_Code'].isin(selected_pcodes)]['Study_ID'].dropna().unique()
        studies = sorted([s for s in studies if s])

        if len(studies) == 0:
            return [{'label': 'No studies found', 'value': 'NONE'}], []

        # Organize studies by phase buckets
        phase_buckets = {
            'Phase 1': [],
            'Phase 2': [],
            'Phase 3': [],
            'Phase 4': [],
            'Other': []}

        for study in studies:
            # Get P_Code for this study
            study_pcode = core_df[core_df['Study_ID'] == study]['P_Code'].iloc[0] if len(core_df[core_df['Study_ID'] == study]) > 0 else None

            # Filter by CDU if specified - use map_df (FULL_MAP) CDU via P_Code
            if cdu_selections and 'ALL' not in cdu_selections and study_pcode:
                study_program = map_df[map_df['P_Code'] == study_pcode]
                if not study_program.empty:
                    study_cdu = study_program.iloc[0]['CDU']
                    if not pd.notna(study_cdu) or study_cdu not in cdu_selections:
                        continue  # Skip this study if it doesn't match CDU filter

            phase = get_study_phase(study, study_daily_df, pcode=study_pcode, map_df=map_df)
            display_name = get_display_name(study, study_daily_df, pcode=study_pcode, map_df=map_df)
            phase_buckets[phase].append((display_name, study))

        # Build dropdown options with phase grouping
        options = [{'label': 'ALL STUDIES', 'value': 'ALL'}]

        # Add phase bucket groups
        phase_order = ['Phase 1', 'Phase 2', 'Phase 3', 'Phase 4', 'Other']

        for phase in phase_order:
            if len(phase_buckets[phase]) > 0:
                # Add phase bucket group selector
                options.append({
                    'label': f'{phase} (All) ----------',
                    'value': f'PHASE_{phase.replace(" ", "_").upper()}'
                })
                # Add individual studies in this phase
                for display_name, study_value in sorted(phase_buckets[phase]):
                    options.append({
                        'label': f'    {display_name}',
                        'value': study_value
                    })

        # Always default to ALL STUDIES
        return options, ['ALL']

    @app.callback(
        [Output('bsb-cdu-dropdown', 'options'),
         Output('bsb-cdu-dropdown', 'value')],
        Input('tabs', 'active_tab'))
    
    def update_bsb_cdu_dropdown(active_tab):
        """Populate CDU dropdown with unique CDUs from Study Daily Report."""
        if study_daily_df.empty:
            return [{'label': 'All CDUs', 'value': 'ALL'}], 'ALL'

        cdus = study_daily_df['CDU'].dropna().unique()
        cdus = sorted([c for c in cdus if c and c != 'Not Assigned'])

        options = [{'label': 'All CDUs', 'value': 'ALL'}] + [{'label': cdu, 'value': cdu} for cdu in cdus]
        return options, 'ALL'

    # UPDATE CDU ON SELECTION =======================
    @app.callback(
        [Output('bsb-cdu-dropdown', 'options'),
         Output('bsb-cdu-dropdown', 'value')],
        Input('tabs', 'active_tab'),
        State('bsb-cdu-dropdown', 'value'))

    def update_bsb_cdus(active_tab, current_value):
        # CDU dropdown is completely independent - shows ALL CDUs from map_df (FULL_MAP)
        # Only populate when on bsb tab
        if active_tab != 'bsb':
            return no_update, no_update

        # Get ALL unique CDUs from map_df (FULL_MAP)
        cdus = map_df['CDU'].dropna().unique()
        cdus = sorted([c for c in cdus if c and c != 'Not Assigned'])

        if len(cdus) == 0:
            return [{'label': 'ALL CDUs', 'value': 'ALL'}], ['ALL']

        options = [{'label': 'ALL CDUs', 'value': 'ALL'}]
        for cdu in cdus:
            options.append({'label': cdu, 'value': cdu})

        # Default to ALL
        return options, ['ALL']

    # FILTER PROGRAMS BY CDU SELECTION =======================
    @app.callback(
        [Output('bsb-project-dropdown', 'options'),
         Output('bsb-project-dropdown', 'value')],
        Input('bsb-cdu-dropdown', 'value'),
        State('bsb-project-dropdown', 'value'))

    def filter_bsb_programs_by_cdu(cdu_selections, current_value):
        # Filter programs by CDU from map_df (FULL_MAP)
        if not cdu_selections or 'ALL' in cdu_selections:
            # Show all programs
            program_options = [{'label': 'ALL PROGRAMS', 'value': 'ALL'}]
            if map_df is not None and not map_df.empty:
                programs = map_df[['P_Code', 'Primary']].drop_duplicates()
                programs = programs[programs['P_Code'].notna() & programs['Primary'].notna()].sort_values('Primary')
                for _, row in programs.iterrows():
                    program_options.append({'label': row['Primary'], 'value': row['P_Code']})
            else:
                programs = core_df[['P_Code', 'Program_Name']].drop_duplicates()
                programs = programs[programs['P_Code'].notna()].sort_values('Program_Name')
                for _, row in programs.iterrows():
                    program_options.append({'label': row['Program_Name'], 'value': row['P_Code']})

            if not current_value or len(current_value) == 0:
                return program_options, ['ALL']
            return program_options, no_update

        # Filter programs by CDU using map_df (FULL_MAP)
        valid_pcodes = []
        if map_df is not None and not map_df.empty:
            for cdu in cdu_selections:
                matching_programs = map_df[map_df['CDU'] == cdu]
                for pcode in matching_programs['P_Code'].dropna().unique():
                    if pcode not in valid_pcodes:
                        valid_pcodes.append(pcode)

        valid_pcodes = sorted(valid_pcodes)

        if len(valid_pcodes) == 0:
            return [{'label': 'No programs found for selected CDU', 'value': 'NONE'}], []

        # Build filtered program options
        program_options = [{'label': 'ALL PROGRAMS', 'value': 'ALL'}]
        if map_df is not None and not map_df.empty:
            for pcode in valid_pcodes:
                map_match = map_df[map_df['P_Code'] == pcode]
                if not map_match.empty and pd.notna(map_match.iloc[0]['Primary']):
                    program_options.append({'label': map_match.iloc[0]['Primary'], 'value': pcode})
                else:
                    program_df = core_df[core_df['P_Code'] == pcode]
                    if not program_df.empty:
                        program_options.append({'label': program_df['Program_Name'].iloc[0], 'value': pcode})
        else:
            for pcode in valid_pcodes:
                program_df = core_df[core_df['P_Code'] == pcode]
                if not program_df.empty:
                    program_options.append({'label': program_df['Program_Name'].iloc[0], 'value': pcode})

        return program_options, ['ALL']

    # LOADING INDICATOR =============================
    @app.callback(
        Output('bsb-loading-indicator', 'children'),
        [Input('bsb-project-dropdown', 'value'),
         Input('bsb-study-dropdown', 'value'),
         Input('bsb-cdu-dropdown', 'value')],
        prevent_initial_call=True
    )
    def show_bsb_loading_indicator(project, study, cdu):
        return '⏳ Updating...'

    @app.callback(
        [Output('bsb-summary-cdu-program', 'children'),
         Output('bsb-summary-cdu', 'children'),
         Output('bsb-summary-indication', 'children'),
         Output('bsb-summary-program', 'children'),
         Output('bsb-summary-project', 'children'),
         Output('bsb-summary-study-count', 'children'),
         Output('bsb-summary-phase', 'children'),
         Output('bsb-summary-status', 'children'),
         Output('bsb-summary-enrollment', 'children'),
         Output('bsb-loading-indicator', 'children', allow_duplicate=True)],
        [Input('bsb-project-dropdown', 'value'),
         Input('bsb-study-dropdown', 'value'),
         Input('bsb-cdu-dropdown', 'value')],
        prevent_initial_call=True)

    def update_bsb_summary(project_codes, study_ids, cdu_selections):
        """Update the BSB page summary card."""
        cdu_program = '—'
        cdu = '—'
        indication = '—'
        program_name = 'None selected'
        project_text = 'None selected'
        study_count = '0'
        program_bucket = None
        phase = '—'
        status = '—'
        enrollment = '—'

        # Show Study/Code dropdown selection
        if study_ids and 'ALL' in study_ids:
            project_text = 'ALL STUDIES'
        elif study_ids and len(study_ids) > 0:
            # Show actual dropdown selection (even if phase groups or non-studies)
            study_display_names = [get_display_name(s, study_daily_df) if not s.startswith('PHASE_') else s.replace('PHASE_', '').replace('_', ' ') for s in study_ids]
            if len(study_display_names) == 1:
                project_text = study_display_names[0]
            else:
                project_text = f"{len(study_display_names)} Selected: {', '.join(study_display_names)}"

        if not project_codes or len(project_codes) == 0:
            # If no program selected, show CDU dropdown selection
            if cdu_selections and 'ALL' in cdu_selections:
                cdu = 'ALL CDUs'
            elif cdu_selections and len(cdu_selections) == 1:
                cdu = cdu_selections[0]
            elif cdu_selections and len(cdu_selections) > 1:
                cdu = f"{len(cdu_selections)} CDUs: {', '.join(cdu_selections)}"
            return cdu_program, cdu, indication, program_name, project_text, study_count, phase, status, enrollment, ''

        # Get list of P_Codes from selection
        if 'ALL' in project_codes:
            selected_pcodes = core_df['P_Code'].dropna().unique().tolist()
            # Show "ALL PROGRAMS" when ALL is selected
            program_name = 'ALL PROGRAMS'
            # When ALL programs, use CDU dropdown selection
            if cdu_selections and 'ALL' in cdu_selections:
                cdu = 'ALL CDUs'
            elif cdu_selections and len(cdu_selections) == 1:
                cdu = cdu_selections[0]
            elif cdu_selections and len(cdu_selections) > 1:
                cdu = f"{len(cdu_selections)} CDUs: {', '.join(cdu_selections)}"
        else:
            selected_pcodes = [pc for pc in project_codes if pc != 'ALL']

            if len(selected_pcodes) == 0:
                return cdu_program, cdu, indication, program_name, project_text, study_count, phase, status, enrollment, ''

            if len(selected_pcodes) == 1:
                project_code = selected_pcodes[0]

                # Get Program Display Name and CDU from map_df
                map_match = map_df[map_df['P_Code'] == project_code]
                if not map_match.empty:
                    if pd.notna(map_match.iloc[0]['Primary']):
                        program_name = map_match.iloc[0]['Primary']
                    program_bucket = map_match.iloc[0]['Program'] if pd.notna(map_match.iloc[0]['Program']) else None
                    # Get CDU from map_df for this program
                    cdu_val = map_match.iloc[0]['CDU'] if pd.notna(map_match.iloc[0]['CDU']) else None
                    if cdu_val:
                        cdu = cdu_val
                        if program_bucket:
                            cdu_program = f'{cdu_val} - {program_bucket}'
                        else:
                            cdu_program = cdu_val
                else:
                    program_df = core_df[core_df['P_Code'] == project_code]
                    if not program_df.empty:
                        program_name = program_df['Program_Name'].iloc[0]
            else:
                # Multiple programs selected - show display names with count
                program_names = []
                for pcode in selected_pcodes:
                    map_match = map_df[map_df['P_Code'] == pcode]
                    if not map_match.empty and pd.notna(map_match.iloc[0]['Primary']):
                        program_names.append(map_match.iloc[0]['Primary'])
                    else:
                        program_df = core_df[core_df['P_Code'] == pcode]
                        if not program_df.empty:
                            program_names.append(program_df['Program_Name'].iloc[0])
                program_name = f"{len(program_names)} Programs: {', '.join(program_names)}" if program_names else 'Multiple programs'

        if 'ALL' in project_codes or len(selected_pcodes) == 1:
            if 'ALL' in project_codes:
                all_pcodes = selected_pcodes
                project_code = None
            else:
                project_code = selected_pcodes[0]
                all_pcodes = [project_code]

            if study_ids and len(study_ids) > 0:
                if 'ALL' in study_ids:
                    # Show "ALL STUDIES" when ALL is selected
                    project_text = 'ALL STUDIES'

                    # Count only actual studies (YYYY-ZZZZ with no letters)
                    if project_code:
                        all_study_ids = core_df[core_df['P_Code'] == project_code]['Study_ID'].dropna().unique()
                    else:
                        all_study_ids = core_df[core_df['P_Code'].isin(all_pcodes)]['Study_ID'].dropna().unique()
                    all_studies = [s for s in all_study_ids if is_actual_study(s)]
                    study_count = str(len(all_studies))

                    for study_id in all_studies:
                        match = re.match(r'\d{4}[-.](\d{3,4})$', str(study_id))
                        if match:
                            study_short = match.group(1)
                            matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]
                            if not matching_row.empty:
                                cdu = matching_row.iloc[0]['CDU'] if pd.notna(matching_row.iloc[0]['CDU']) else None
                                if cdu and program_bucket:
                                    cdu_program = f'{cdu} - {program_bucket}'
                                elif cdu:
                                    cdu_program = cdu
                                break
                else:
                    # Get actual studies from selection
                    actual_studies = [s for s in study_ids if is_actual_study(s)]
                    study_count = str(len(actual_studies))

                    if len(actual_studies) == 1:
                        # Single study - get all study info
                        match = re.match(r'\d{4}[-.](\d{3,4})$', str(actual_studies[0]))
                        if match:
                            study_short = match.group(1)
                            matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]
                            if not matching_row.empty:
                                indication = matching_row.iloc[0]['Indication'] if pd.notna(matching_row.iloc[0]['Indication']) else '—'
                                phase = matching_row.iloc[0]['Phase'] if pd.notna(matching_row.iloc[0]['Phase']) else '—'
                                status = matching_row.iloc[0]['Study Status (COMPASS)'] if pd.notna(matching_row.iloc[0]['Study Status (COMPASS)']) else '—'

                                # Enrollment: Actual/Planned
                                enroll_actual = matching_row.iloc[0]['# Enrollment (Actual)'] if pd.notna(matching_row.iloc[0]['# Enrollment (Actual)']) else 0
                                enroll_planned = matching_row.iloc[0]['# Enrollment (Planned)'] if pd.notna(matching_row.iloc[0]['# Enrollment (Planned)']) else 0
                                enrollment = f"{int(enroll_actual)}/{int(enroll_planned)}"
                    else:
                        # Multiple studies - show each on its own row
                        study_phases = []
                        study_statuses = []
                        study_enrollments = []

                        for study in actual_studies:
                            match = re.match(r'\d{4}[-.](\d{3,4})$', str(study))
                            if match:
                                study_short = match.group(1)
                                matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]
                                if not matching_row.empty:
                                    study_name = get_display_name(study, study_daily_df)
                                    study_phase = matching_row.iloc[0]['Phase'] if pd.notna(matching_row.iloc[0]['Phase']) else '—'
                                    study_status = matching_row.iloc[0]['Study Status (COMPASS)'] if pd.notna(matching_row.iloc[0]['Study Status (COMPASS)']) else '—'
                                    enroll_actual = matching_row.iloc[0]['# Enrollment (Actual)'] if pd.notna(matching_row.iloc[0]['# Enrollment (Actual)']) else 0
                                    enroll_planned = matching_row.iloc[0]['# Enrollment (Planned)'] if pd.notna(matching_row.iloc[0]['# Enrollment (Planned)']) else 0

                                    study_phases.append(f"{study_name}: {study_phase}")
                                    study_statuses.append(f"{study_name}: {study_status}")
                                    study_enrollments.append(f"{study_name}: {int(enroll_actual)}/{int(enroll_planned)}")

                        indication = 'Multiple'
                        phase = html.Div([html.Div(p, style={'marginBottom': '2px'}) for p in study_phases]) if study_phases else '—'
                        status = html.Div([html.Div(s, style={'marginBottom': '2px'}) for s in study_statuses]) if study_statuses else '—'
                        enrollment = html.Div([html.Div(e, style={'marginBottom': '2px'}) for e in study_enrollments]) if study_enrollments else '—'

        return cdu_program, cdu, indication, program_name, project_text, study_count, phase, status, enrollment, ''

    @app.callback(
        [Output('bsb-gauge-combined', 'figure'),
         Output('bsb-total-budget', 'children'),
         Output('bsb-committed-total', 'children'),
         Output('bsb-actuals-spent', 'children'),
         Output('bsb-committed-remaining', 'children'),
         Output('bsb-total-forecasted', 'children'),
         Output('bsb-projected-total', 'children'),
         Output('bsb-delta', 'children'),
         Output('bsb-delta', 'style'),
         Output('bsb-remaining', 'children'),
         Output('bsb-burn-rate', 'children'),
         Output('bsb-overrun', 'children'),
         Output('bsb-quarters-left', 'children'),
         Output('bsb-stacked-bar', 'figure'),
         Output('bsb-loading-indicator', 'children', allow_duplicate=True)],
        [Input('bsb-project-dropdown', 'value'),
         Input('bsb-study-dropdown', 'value'),
         Input('bsb-cdu-dropdown', 'value'),
         Input('theme-store', 'data')],
        prevent_initial_call=True)

    def update_bsb_dashboard(project_codes, study_ids, cdu_selections, theme):
        """Update BSB dashboard with budget metrics and charts - COMPLETE LOGIC FROM APOFIS_DEV.py"""
        # Always show ALL accounts
        account_codes = ['ALL']
        # Get colors for current theme
        colors = get_colors(theme)

        # Default blank figures
        blank_gauge = go.Figure()
        blank_gauge.update_layout(
            plot_bgcolor=colors['surface-2'],
            paper_bgcolor=colors['surface-2'],
            font=dict(color=colors['ink-body']))

        blank_bar = go.Figure()
        blank_bar.update_layout(
            plot_bgcolor=colors['surface-2'],
            paper_bgcolor=colors['surface-2'],
            font=dict(color=colors['ink-body']),
            barmode='group',
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor=colors['line-soft']),
            height=550,
            margin=dict(l=80, r=40, t=100, b=150))

        # Default values
        total_budget = '$0'
        committed_total = '$0'
        actuals_spent = '$0 realized (0%)'
        committed_remaining = '$0 remaining'
        forecast_only = '$0'
        projected_total = '$0'
        delta_text = '▼ $0 (0.0%)'
        delta_style = {'fontSize': '16px', 'fontWeight': 'bold', 'textAlign': 'center', 'color': colors['brand-green']}
        remaining = '$0'
        burn_rate = '$0/qtr'
        overrun = '0%'
        quarters_left = 'N/A'

        if not project_codes or len(project_codes) == 0:
            return blank_gauge, total_budget, committed_total, actuals_spent, committed_remaining, forecast_only, projected_total, delta_text, delta_style, remaining, burn_rate, overrun, quarters_left, blank_bar, ''

        # Get list of P_Codes from selection
        if 'ALL' in project_codes:
            selected_pcodes = core_df['P_Code'].dropna().unique().tolist()
        else:
            selected_pcodes = [pc for pc in project_codes if pc != 'ALL']

        if len(selected_pcodes) == 0:
            return blank_gauge, total_budget, committed_total, actuals_spent, committed_remaining, forecast_only, projected_total, delta_text, delta_style, remaining, burn_rate, overrun, quarters_left, blank_bar, ''

        # Filter data
        filtered = core_df[core_df['P_Code'].isin(selected_pcodes)].copy()
        if study_ids and len(study_ids) > 0 and 'ALL' not in study_ids:
            expanded_study_ids = expand_phase_groups(study_ids, selected_pcodes)
            filtered = filtered[filtered['Study_ID'].isin(expanded_study_ids)]

        # Handle therapeutic area filter (multi-select)
        if cdu_selections and 'ALL' not in cdu_selections and len(cdu_selections) > 0:
            # Filter to only studies matching selected TAs
            study_ids_with_ta = []
            unique_study_ids = filtered['Study_ID'].dropna().unique()
            for study_id in unique_study_ids:
                match = re.match(r'\d{4}[-.](\d{3,4})$', str(study_id))
                if match:
                    study_short = match.group(1)
                    matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]
                    if not matching_row.empty:
                        study_ta = matching_row.iloc[0]['CDU']
                        if pd.notna(study_ta) and study_ta in cdu_selections:
                            study_ids_with_ta.append(study_id)
            # Apply TA filter to dataframe
            filtered = filtered[filtered['Study_ID'].isin(study_ids_with_ta)]

        if account_codes and len(account_codes) > 0 and 'ALL' not in account_codes:
            if 'PF_GROUP' in account_codes:
                pf_account_list = [acc for acc in filtered['Account_Clean'].unique() if is_program_finance_account(acc)]
                filtered = filtered[filtered['Account_Clean'].isin(pf_account_list)]
            else:
                filtered = filtered[filtered['Account_Clean'].isin(account_codes)]

        # Calculate metrics
        # NOTE: core_df already has forecast data merged in (both shared and unique quarters)
        # So we can calculate everything from filtered (the already-merged core_df subset)
        actuals_total = filtered[quarter_cols].sum().sum()

        # Get ONLY the future forecast quarters (quarters that exist in forecast_quarter_cols but not in quarter_cols)
        future_forecast_quarters = [q for q in forecast_quarter_cols if q not in quarter_cols]

        # Sum the future forecast quarters from filtered (which has them merged in already)
        forecast_total = filtered[future_forecast_quarters].sum().sum() if len(future_forecast_quarters) > 0 else 0

        # Total projected = actuals (historical) + forecast (future)
        total_projected_amt = actuals_total + forecast_total

        # ========== GET REAL BUDGET FROM BSB DATA ==========
        # Filter BSB data by P_Code
        bsb_filtered = bsb_df[bsb_df['P_Code'].isin(selected_pcodes)].copy()

        # Filter by study if specific studies selected
        if study_ids and len(study_ids) > 0 and 'ALL' not in study_ids:
            expanded_study_ids = expand_phase_groups(study_ids, selected_pcodes)
            bsb_filtered = bsb_filtered[bsb_filtered['Study_ID'].isin(expanded_study_ids)]

        # Filter by therapeutic area if specific TAs selected
        if cdu_selections and 'ALL' not in cdu_selections and len(cdu_selections) > 0:
            study_ids_with_ta = []
            unique_study_ids = bsb_filtered['Study_ID'].dropna().unique()
            for study_id in unique_study_ids:
                match = re.match(r'\d{4}[-.](\d{3,4})$', str(study_id))
                if match:
                    study_short = match.group(1)
                    matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]
                    if not matching_row.empty:
                        study_ta = matching_row.iloc[0]['CDU']
                        if pd.notna(study_ta) and study_ta in cdu_selections:
                            study_ids_with_ta.append(study_id)
            bsb_filtered = bsb_filtered[bsb_filtered['Study_ID'].isin(study_ids_with_ta)]

        # Filter by accounts if specific accounts selected
        if account_codes and len(account_codes) > 0 and 'ALL' not in account_codes:
            if 'PF_GROUP' in account_codes:
                # Filter to Program Finance accounts only
                pf_account_codes = [extract_account_code(acc) for acc in filtered['Account_Clean'].unique() if is_program_finance_account(acc)]
                bsb_filtered = bsb_filtered[bsb_filtered['Account_Code'].isin(pf_account_codes)]
            else:
                # Filter to specific selected accounts
                selected_account_codes = [extract_account_code(acc) for acc in account_codes]
                bsb_filtered = bsb_filtered[bsb_filtered['Account_Code'].isin(selected_account_codes)]

        # Calculate total budget from BSB data
        budget = bsb_filtered['Budget'].sum()

        # If no budget found (edge case), use a small placeholder to avoid division by zero
        if budget == 0:
            budget = 1  # Placeholder to avoid math errors

        # ========== GET PO COMMITMENTS FROM PO DATA ==========
        # Filter PO data by P_Code
        po_filtered = po_df[po_df['P_Code'].isin(selected_pcodes)].copy()

        # Filter by study if specific studies selected
        # Match PO data by Project_Code_Full (e.g., "P570OD00C3831") to actuals Project_Code_Full
        if study_ids and len(study_ids) > 0 and 'ALL' not in study_ids:
            # Get the Project_Code_Full values for the selected studies from filtered actuals
            study_project_codes = filtered['Project_Code_Full'].unique()
            po_filtered = po_filtered[po_filtered['Project_Code_Full'].isin(study_project_codes)]

        # Filter by accounts if specific accounts selected
        # NOTE: We match by account CODE (e.g., "A80070"), not by full Account_Clean string,
        # because PO and actuals have different capitalization in descriptions
        if account_codes and len(account_codes) > 0 and 'ALL' not in account_codes:
            if 'PF_GROUP' in account_codes:
                # Filter to Program Finance accounts only
                pf_account_codes = [extract_account_code(acc) for acc in filtered['Account_Clean'].unique() if is_program_finance_account(acc)]
                # Extract account codes from PO data for matching
                po_filtered['Account_Code_Match'] = po_filtered['Account_Clean'].apply(extract_account_code)
                po_filtered = po_filtered[po_filtered['Account_Code_Match'].isin(pf_account_codes)]
            else:
                # Filter to specific selected accounts by extracting codes
                selected_account_codes = [extract_account_code(acc) for acc in account_codes]
                # Extract account codes from PO data for matching
                po_filtered['Account_Code_Match'] = po_filtered['Account_Clean'].apply(extract_account_code)
                po_filtered = po_filtered[po_filtered['Account_Code_Match'].isin(selected_account_codes)]

        # Calculate total PO commitment
        po_commitment = po_filtered['Commitment'].sum()

        # Calculate Committed Spend = Actuals + PO Commitments
        committed_spend_total = actuals_total + po_commitment
        actuals_pct_of_committed = (actuals_total / committed_spend_total * 100) if committed_spend_total > 0 else 0

        # Format the main numbers
        total_budget = f'${budget:,.0f}'
        # Committed Spend - split into two lines
        committed_total = f'${committed_spend_total:,.0f}'
        actuals_spent = f'${actuals_total:,.0f} realized ({actuals_pct_of_committed:.0f}%)'
        # Calculate remaining from committed spend (Committed - Actuals = PO commitments remaining)
        committed_remaining_amt = committed_spend_total - actuals_total
        committed_remaining = f'${committed_remaining_amt:,.0f} remaining'
        forecast_only = f'${forecast_total:,.0f}'  # ONLY future forecast, not including actuals
        projected_total = f'${total_projected_amt:,.0f}'  # Actuals + future forecast

        # Calculate delta vs budget
        delta_amt = total_projected_amt - budget
        delta_pct = (delta_amt / budget * 100) if budget > 0 else 0
        delta_color = '#EF553B' if delta_amt > 0 else colors['brand-green']  # Red if over, green if under
        delta_symbol = '▲' if delta_amt > 0 else '▼'
        delta_text = f'{delta_symbol} ${abs(delta_amt):,.0f} ({abs(delta_pct):.1f}%)'
        delta_style = {'fontSize': '16px', 'fontWeight': 'bold', 'textAlign': 'center', 'color': delta_color}

        # Calculations
        budget_used_pct = (actuals_total / budget * 100) if budget > 0 else 0
        committed_spend_pct = (committed_spend_total / budget * 100) if budget > 0 else 0
        forecast_vs_budget_pct = (total_projected_amt / budget * 100) if budget > 0 else 0
        remaining_amt = budget - committed_spend_total  # Remaining after committed spend
        remaining = f'${remaining_amt:,.0f}'

        # Burn rate (average per quarter)
        num_quarters_with_data = sum(1 for q in quarter_cols if filtered[q].sum() > 0)
        avg_burn = actuals_total / num_quarters_with_data if num_quarters_with_data > 0 else 0
        burn_rate = f'${avg_burn:,.0f}/qtr'

        # Overrun risk
        if total_projected_amt > budget:
            overrun_amt = total_projected_amt - budget
            overrun_pct = (overrun_amt / budget * 100) if budget > 0 else 0
            overrun = f'+{overrun_pct:.0f}%'
            overrun_color = colors['brand-cyan'] if overrun_pct < 10 else '#EF553B'
        else:
            overrun = '0%'
            overrun_color = colors['brand-green']

        # Quarters to depletion
        if avg_burn > 0 and remaining_amt > 0:
            qtrs_left = remaining_amt / avg_burn
            quarters_left = f'{qtrs_left:.1f} qtrs'
        else:
            quarters_left = 'N/A'

        # COMBINED GAUGE: Shows Actuals + PO Commitments + Forecast vs Budget on one gauge
        actuals_display = f'${actuals_total/1e9:.2f}B' if actuals_total >= 1e9 else f'${actuals_total/1e6:.1f}M'
        po_display = f'${po_commitment/1e9:.2f}B' if po_commitment >= 1e9 else f'${po_commitment/1e6:.1f}M'
        committed_display = f'${committed_spend_total/1e9:.2f}B' if committed_spend_total >= 1e9 else f'${committed_spend_total/1e6:.1f}M'
        projected_display = f'${total_projected_amt/1e9:.2f}B' if total_projected_amt >= 1e9 else f'${total_projected_amt/1e6:.1f}M'
        budget_display = f'${budget/1e9:.2f}B' if budget >= 1e9 else f'${budget/1e6:.1f}M'

        # Calculate percentages for gauge
        actuals_with_po_pct = committed_spend_pct  # This is the same as committed spend %

        # Create figure with multiple indicators
        combined_gauge = go.Figure()

        # LAYER 1: Main gauge with axis, title, and number display (no bar - transparent)
        combined_gauge.add_trace(go.Indicator(
            mode='gauge+number',
            value=committed_spend_pct,
            domain={'x': [0, 1], 'y': [0, 1]},
            title={'text': f'Budget Status<br><sub>Committed: {committed_display} (Actuals: {actuals_display}, POs: {po_display}) | Projected: {projected_display} | Budget: {budget_display}</sub>',
                   'font': {'size': 18, 'color': colors['ink-body']}},
            number={'suffix': '% Committed', 'font': {'size': 32, 'color': colors['brand-green']}},
            gauge={
                'axis': {'range': [0, 120], 'tickcolor': colors['ink-soft'], 'tickwidth': 2},
                'bar': {'color': 'rgba(0, 0, 0, 0)', 'thickness': 0},  # Transparent bar - no display
                'bgcolor': 'rgba(0, 0, 0, 0)',
                'steps': [
                    {'range': [0, 50], 'color': 'rgba(0, 255, 156, 0.15)'},
                    {'range': [50, 80], 'color': 'rgba(255, 193, 7, 0.15)'},
                    {'range': [80, 100], 'color': 'rgba(239, 83, 59, 0.15)'},
                    {'range': [100, 120], 'color': 'rgba(239, 83, 59, 0.3)'}],
                'threshold': {
                    'line': {'color': 'white', 'width': 3},
                    'thickness': 0.8,
                    'value': 100}
            }
        ))

        # LAYER 2: Actuals realized bar (solid green)
        combined_gauge.add_trace(go.Indicator(
            mode='gauge',
            value=budget_used_pct,
            domain={'x': [0, 1], 'y': [0, 1]},
            gauge={
                'axis': {'range': [0, 120], 'visible': False},
                'bar': {'color': colors['brand-green'], 'thickness': 0.6},
                'bgcolor': 'rgba(0, 0, 0, 0)'}
        ))

        # LAYER 3: PO commitment overlay (hashed/transparent green extending from actuals to committed)
        if po_commitment > 0:
            combined_gauge.add_trace(go.Indicator(
                mode='gauge',
                value=committed_spend_pct,
                domain={'x': [0, 1], 'y': [0, 1]},
                gauge={
                    'axis': {'range': [0, 120], 'visible': False},
                    'bar': {'color': 'rgba(0, 255, 156, 0.35)', 'thickness': 0.6, 'line': {'color': colors['brand-green'], 'width': 2}},
                    'bgcolor': 'rgba(0, 0, 0, 0)'}
            ))

        # Add forecast indicator (cyan/blue bar overlay)
        combined_gauge.add_trace(go.Indicator(
            mode='gauge',
            value=forecast_vs_budget_pct,
            domain={'x': [0, 1], 'y': [0, 1]},
            gauge={
                'axis': {'range': [0, 120], 'visible': False},
                'bar': {'color': colors['brand-cyan'], 'thickness': 0.3},
                'bgcolor': 'rgba(0, 0, 0, 0)'}
        ))

        combined_gauge.update_layout(
            paper_bgcolor=colors['surface-2'],
            font={'color': colors['ink-body']},
            height=420,
            margin=dict(l=20, r=20, t=50, b=20))

        # GROUPED BAR CHART: Budget vs Actuals vs Forecasted per Account
        bar_fig = go.Figure()

        # Get accounts to display
        if account_codes and 'ALL' not in account_codes and 'PF_GROUP' not in account_codes:
            accounts_to_show = account_codes
        else:
            accounts_to_show = filtered['Account_Clean'].dropna().unique()

        # Sort accounts and limit to top 15 by budget
        accounts_sorted = sorted(accounts_to_show)[:15] if len(accounts_to_show) > 0 else []

        # Prepare data for each account
        account_names = []
        budgets = []
        actuals = []
        forecasts = []
        po_commits = []

        if len(accounts_sorted) > 0:
            for account in accounts_sorted:
                # Get actuals for this account from filtered
                account_actuals = filtered[filtered['Account_Clean'] == account]
                actuals_total = account_actuals[quarter_cols].sum().sum()

                # Get forecast for this account from filtered (future quarters only)
                forecast_total = account_actuals[future_forecast_quarters].sum().sum() if len(future_forecast_quarters) > 0 else 0

                # Get budget for this account from bsb_filtered
                account_code = extract_account_code(account)
                account_budget = bsb_filtered[bsb_filtered['Account_Code'] == account_code]['Budget'].sum()

                # Get PO commitment for this account from po_filtered
                # Match by account CODE, not full string (case-insensitive matching)
                if 'Account_Code_Match' in po_filtered.columns:
                    account_po = po_filtered[po_filtered['Account_Code_Match'] == account_code]['Commitment'].sum()
                else:
                    # Fallback: extract code on the fly
                    po_account_codes = po_filtered['Account_Clean'].apply(extract_account_code)
                    account_po = po_filtered[po_account_codes == account_code]['Commitment'].sum()

                # Add to lists
                account_names.append(account)
                budgets.append(account_budget)
                actuals.append(actuals_total)
                forecasts.append(actuals_total + forecast_total)  # Total forecasted = actuals + future
                po_commits.append(account_po)

            # Add Budget bars (gray, transparent)
            bar_fig.add_trace(go.Bar(
                name='Budget (BSB)',
                x=account_names,
                y=budgets,
                marker=dict(color=colors['ink-soft'], opacity=0.5, line=dict(color=colors['ink-body'], width=1)),
                text=[f'${b/1e6:.1f}M' if b >= 1e6 else f'${b/1e3:.0f}K' for b in budgets],
                textposition='outside',
                textfont=dict(size=10),
                hovertemplate='<b>Budget</b><br>%{y:$,.0f}<extra></extra>'))

            # Add Actuals bars (green)
            bar_fig.add_trace(go.Bar(
                name='Actuals Spent',
                x=account_names,
                y=actuals,
                marker=dict(color=colors['brand-green']),
                text=[f'${a/1e6:.1f}M' if a >= 1e6 else f'${a/1e3:.0f}K' for a in actuals],
                textposition='outside',
                textfont=dict(size=10),
                hovertemplate='<b>Actuals</b><br>%{y:$,.0f}<extra></extra>',
                error_y=dict(
                    type='data',
                    symmetric=False,
                    array=po_commits,  # Extend upward from actuals by PO commitment amount
                    arrayminus=[0] * len(po_commits),  # No extension downward
                    color='rgba(0, 255, 156, 0.6)',
                    thickness=3,
                    width=8)
            ))

            # Add Forecasted bars (cyan/blue with pattern)
            bar_fig.add_trace(go.Bar(
                name='Total Projected',
                x=account_names,
                y=forecasts,
                marker=dict(
                    color=colors['brand-cyan'],
                    opacity=0.7,
                    line=dict(color=colors['brand-cyan'], width=2),
                    pattern=dict(shape="/", fgcolor="white", bgcolor=colors['brand-cyan'], size=8, solidity=0.3)
                ),
                text=[f'${f/1e6:.1f}M' if f >= 1e6 else f'${f/1e3:.0f}K' for f in forecasts],
                textposition='outside',
                textfont=dict(size=10),
                hovertemplate='<b>Projected Total</b><br>%{y:$,.0f}<extra></extra>'
            ))

        bar_fig.update_layout(
            barmode='group',  # Side-by-side bars
            plot_bgcolor=colors['surface-2'],
            paper_bgcolor=colors['surface-2'],
            font=dict(color=colors['ink-body']),
            xaxis=dict(
                title='Account',
                showgrid=False,
                tickangle=-45,
                tickfont=dict(size=9)),
            yaxis=dict(
                title='Amount ($)',
                showgrid=True,
                gridcolor=colors['line-soft']),
            legend=dict(
                orientation='h',
                yanchor='bottom',
                y=1.02,
                xanchor='center',
                x=0.5,
                bgcolor=colors['surface-2'],
                bordercolor=colors['line-soft'],
                borderwidth=1,
                font=dict(color=colors['ink-body'])),
            height=550,
            margin=dict(l=80, r=40, t=100, b=150)
        )

        return combined_gauge, total_budget, committed_total, actuals_spent, committed_remaining, forecast_only, projected_total, delta_text, delta_style, remaining, burn_rate, overrun, quarters_left, bar_fig, ''

    # ==================== BSB DATA TABLE MODAL ====================

    @app.callback(
        Output('bsb-data-modal', 'is_open'),
        [Input('bsb-view-data-btn', 'n_clicks'),
         Input('bsb-close-modal-btn', 'n_clicks')],
        State('bsb-data-modal', 'is_open'))
    
    def toggle_bsb_modal(n_open, n_close, is_open):
        """Toggle the BSB data table modal."""
        if n_open or n_close:
            return not is_open
        return is_open

    @app.callback(
        Output('bsb-data-table-container', 'children'),
        [Input('bsb-project-dropdown', 'value'),
         Input('bsb-study-dropdown', 'value'),
         Input('bsb-cdu-dropdown', 'value'),
         Input('theme-store', 'data')])

    def update_bsb_data_table(project_codes, study_ids, cdu_selections, theme):
        """Generate data table showing filtered actuals and budgets."""
        # Always show ALL accounts
        account_codes = ['ALL']
        # Get colors for current theme
        colors = get_colors(theme)

        if not project_codes or len(project_codes) == 0:
            return html.P('Select a program to view data.', style={'color': colors['ink-soft']})

        # Get list of P_Codes from selection
        if 'ALL' in project_codes:
            selected_pcodes = core_df['P_Code'].dropna().unique().tolist()
        else:
            selected_pcodes = [pc for pc in project_codes if pc != 'ALL']

        # Apply same filters as BSB dashboard
        filtered = core_df[core_df['P_Code'].isin(selected_pcodes)].copy()
        if study_ids and len(study_ids) > 0 and 'ALL' not in study_ids:
            expanded_study_ids = expand_phase_groups(study_ids, selected_pcodes)
            filtered = filtered[filtered['Study_ID'].isin(expanded_study_ids)]

        # Filter by therapeutic area if specific TAs selected
        if cdu_selections and 'ALL' not in cdu_selections and len(cdu_selections) > 0:
            study_ids_with_ta = []
            unique_study_ids = filtered['Study_ID'].dropna().unique()
            for study_id in unique_study_ids:
                match = re.match(r'\d{4}[-.](\d{3,4})$', str(study_id))
                if match:
                    study_short = match.group(1)
                    matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]
                    if not matching_row.empty:
                        study_ta = matching_row.iloc[0]['CDU']
                        if pd.notna(study_ta) and study_ta in cdu_selections:
                            study_ids_with_ta.append(study_id)
            filtered = filtered[filtered['Study_ID'].isin(study_ids_with_ta)]

        if account_codes and len(account_codes) > 0 and 'ALL' not in account_codes:
            if 'PF_GROUP' in account_codes:
                pf_account_list = [acc for acc in filtered['Account_Clean'].unique() if is_program_finance_account(acc)]
                filtered = filtered[filtered['Account_Clean'].isin(pf_account_list)]
            else:
                filtered = filtered[filtered['Account_Clean'].isin(account_codes)]

        # Get BSB data with same filters
        bsb_filtered = bsb_df[bsb_df['P_Code'].isin(selected_pcodes)].copy()
        if study_ids and len(study_ids) > 0 and 'ALL' not in study_ids:
            expanded_study_ids = expand_phase_groups(study_ids, selected_pcodes)
            bsb_filtered = bsb_filtered[bsb_filtered['Study_ID'].isin(expanded_study_ids)]

        # Filter by therapeutic area if specific TAs selected
        if cdu_selections and 'ALL' not in cdu_selections and len(cdu_selections) > 0:
            study_ids_with_ta = []
            unique_study_ids = bsb_filtered['Study_ID'].dropna().unique()
            for study_id in unique_study_ids:
                match = re.match(r'\d{4}[-.](\d{3,4})$', str(study_id))
                if match:
                    study_short = match.group(1)
                    matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]
                    if not matching_row.empty:
                        study_ta = matching_row.iloc[0]['CDU']
                        if pd.notna(study_ta) and study_ta in cdu_selections:
                            study_ids_with_ta.append(study_id)
            bsb_filtered = bsb_filtered[bsb_filtered['Study_ID'].isin(study_ids_with_ta)]

        if account_codes and len(account_codes) > 0 and 'ALL' not in account_codes:
            if 'PF_GROUP' in account_codes:
                pf_account_codes = [extract_account_code(acc) for acc in filtered['Account_Clean'].unique() if is_program_finance_account(acc)]
                bsb_filtered = bsb_filtered[bsb_filtered['Account_Code'].isin(pf_account_codes)]
            else:
                selected_account_codes = [extract_account_code(acc) for acc in account_codes]
                bsb_filtered = bsb_filtered[bsb_filtered['Account_Code'].isin(selected_account_codes)]

        # Calculate actuals per account
        future_forecast_quarters = [q for q in forecast_quarter_cols if q not in quarter_cols]
        summary_data = []

        for account in filtered['Account_Clean'].dropna().unique():
            account_data = filtered[filtered['Account_Clean'] == account]
            actuals = account_data[quarter_cols].sum().sum()
            forecast = account_data[future_forecast_quarters].sum().sum() if len(future_forecast_quarters) > 0 else 0
            projected = actuals + forecast

            # Get budget
            account_code = extract_account_code(account)
            budget = bsb_filtered[bsb_filtered['Account_Code'] == account_code]['Budget'].sum()

            summary_data.append({
                'Account': account,
                'Actuals': f'${actuals:,.0f}',
                'Forecast': f'${forecast:,.0f}',
                'Projected Total': f'${projected:,.0f}',
                'BSB Budget': f'${budget:,.0f}',
                'Delta': f'${projected - budget:,.0f}',
                '% of Budget': f'{(projected/budget*100):.1f}%' if budget > 0 else 'N/A'
            })

        # Create DataFrame for display
        import pandas as pd
        df_summary = pd.DataFrame(summary_data)

        # Add totals row
        total_actuals = filtered[quarter_cols].sum().sum()
        total_forecast = filtered[future_forecast_quarters].sum().sum() if len(future_forecast_quarters) > 0 else 0
        total_projected = total_actuals + total_forecast
        total_budget = bsb_filtered['Budget'].sum()

        totals_row = pd.DataFrame([{
            'Account': 'TOTAL',
            'Actuals': f'${total_actuals:,.0f}',
            'Forecast': f'${total_forecast:,.0f}',
            'Projected Total': f'${total_projected:,.0f}',
            'BSB Budget': f'${total_budget:,.0f}',
            'Delta': f'${total_projected - total_budget:,.0f}',
            '% of Budget': f'{(total_projected/total_budget*100):.1f}%' if total_budget > 0 else 'N/A'
        }])

        df_summary = pd.concat([df_summary, totals_row], ignore_index=True)

        # Create Dash DataTable
        from dash import dash_table
        return dash_table.DataTable(
            data=df_summary.to_dict('records'),
            columns=[{'name': col, 'id': col} for col in df_summary.columns],
            style_cell={'textAlign': 'left',
                        'padding': '10px',
                        'backgroundColor': colors['surface-2'],
                        'color': colors['ink-body'],
                        'border': f'1px solid {colors["line-soft"]}'},
            style_header={'backgroundColor': colors['surface-0'],
                          'fontWeight': 'bold',
                          'color': colors['ink-strong'],
                          'border': f'1px solid {colors["line-soft"]}'},
            style_data_conditional=[{'if': {'filter_query': '{Account} = "TOTAL"'},
                                     'fontWeight': 'bold',
                                     'backgroundColor': colors['surface-0'],
                                     'color': colors['brand-cyan']}],
            page_size=20
        )

    # ==================== FILTER SYNC CALLBACKS ====================

    @app.callback(
        [Output('forecast-project-dropdown', 'value'),
         Output('forecast-study-dropdown', 'value')],
        Input('actuals-send-filters-btn', 'n_clicks'),
        [State('actuals-project-dropdown', 'value'),
         State('actuals-study-dropdown', 'value')],
        prevent_initial_call=True)

    def send_actuals_to_forecast(n_clicks, project, study):
        """Copy filter selections from Actuals to Forecast page (no account filter on Forecast)."""
        return project, study

    @app.callback(
        [Output('actuals-project-dropdown', 'value'),
         Output('actuals-study-dropdown', 'value')],
        Input('forecast-send-filters-btn', 'n_clicks'),
        [State('forecast-project-dropdown', 'value'),
         State('forecast-study-dropdown', 'value')],
        prevent_initial_call=True)

    def send_forecast_to_actuals(n_clicks, project, study):
        """Copy filter selections from Forecast to Actuals page (no account dropdown on any tab)."""
        return project, study

    # ==================== FORECAST GENERATION CALLBACKS ====================

    @app.callback(
        [Output('fpfv-date', 'date'),
         Output('fpfd-date', 'date'),
         Output('lpfd-date', 'date'),
         Output('lplv-date', 'date'),
         Output('dbl-date', 'date')],
        Input('forecast-study-dropdown', 'value'),
        prevent_initial_call=True)
    
    def populate_milestone_dates(selected_studies):
        """Auto-populate milestone dates when a single study is selected."""
        # Return None for all dates if no selection or multiple studies
        if not selected_studies or len(selected_studies) != 1 or 'ALL' in selected_studies:
            return None, None, None, None, None

        study_id = selected_studies[0]

        # Extract study short number to match with study_daily_df
        match = re.match(r'\d{4}[-.](\d{3,4})$', str(study_id))
        if not match:
            return None, None, None, None, None

        study_short = match.group(1)
        matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]

        if matching_row.empty:
            return None, None, None, None, None

        row = matching_row.iloc[0]

        # Extract milestone dates
        fpfv = row.get('FPFV')
        fpfd = row.get('FPFD')
        lpfd = row.get('LPFD')
        lplv = row.get('LPLV')
        dbl = row.get('DBL')

        # Convert to date strings or None
        def to_date_str(date_val):
            if pd.isna(date_val):
                return None
            return date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else None

        return to_date_str(fpfv), to_date_str(fpfd), to_date_str(lpfd), to_date_str(lplv), to_date_str(dbl)

    @app.callback(
        [Output('fpfv-date', 'date', allow_duplicate=True),
         Output('fpfd-date', 'date', allow_duplicate=True),
         Output('lpfd-date', 'date', allow_duplicate=True),
         Output('lplv-date', 'date', allow_duplicate=True),
         Output('dbl-date', 'date', allow_duplicate=True)],
        Input('reset-milestones-btn', 'n_clicks'),
        State('forecast-study-dropdown', 'value'),
        prevent_initial_call=True)
    
    def reset_milestone_dates(n_clicks, selected_studies):
        """Reset milestone dates to original values from Study Daily Report."""
        if not n_clicks:
            return None, None, None, None, None

        # Same logic as populate_milestone_dates
        if not selected_studies or len(selected_studies) != 1 or 'ALL' in selected_studies:
            return None, None, None, None, None

        study_id = selected_studies[0]
        match = re.match(r'\d{4}[-.](\d{3,4})$', str(study_id))
        if not match:
            return None, None, None, None, None

        study_short = match.group(1)
        matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]

        if matching_row.empty:
            return None, None, None, None, None

        row = matching_row.iloc[0]

        def to_date_str(date_val):
            if pd.isna(date_val):
                return None
            return date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else None

        return to_date_str(row.get('FPFV')), to_date_str(row.get('FPFD')), to_date_str(row.get('LPFD')), to_date_str(row.get('LPLV')), to_date_str(row.get('DBL'))

    @app.callback(
        [Output('forecast-chart', 'figure'),
         Output('forecast-account-table', 'children'),
         Output('custom-forecast-store', 'data'),
         Output('monthly-table-store', 'data'),
         Output('forecast-loading-indicator', 'children', allow_duplicate=True)],
        Input('generate-forecast-btn', 'n_clicks'),
        [State('forecast-method', 'value'),
         State('show-account-lines-toggle', 'value'),
         State('show-confidence-toggle', 'value'),
         State('show-previous-forecast-toggle', 'value'),
         State('bsb-total-input', 'value'),
         State('fpfv-date', 'date'),
         State('fpfd-date', 'date'),
         State('lpfd-date', 'date'),
         State('lplv-date', 'date'),
         State('dbl-date', 'date'),
         State('forecast-project-dropdown', 'value'),
         State('forecast-study-dropdown', 'value'),
         State('bsb-adjustments-store', 'data'),
         State('theme-store', 'data')],
        prevent_initial_call=True)

    def generate_study_forecast(n_clicks, method, show_acct_lines, show_conf, show_prev, bsb_total_override,
                                fpfv, fpfd, lpfd, lplv, dbl, selected_programs, selected_studies, bsb_adjustments, theme):
        """Generate forecast based on selected method, milestones, and filters."""
        # Always use ALL accounts (no account filter dropdown)
        selected_accounts = ['ALL']
        """Generate forecast based on selected method, milestones, and filters."""
        import forecast_logic
        from dash import dash_table

        # Get colors for current theme
        colors = get_colors(theme)

        if not n_clicks:
            return go.Figure(), html.Div("Click 'Generate Forecast' to see results."), None, None, ''

        # Check if we have valid selections
        if not selected_programs or not selected_studies:
            empty_fig = go.Figure()
            empty_fig.update_layout(
                plot_bgcolor=colors['surface-2'],
                paper_bgcolor=colors['surface-2'],
                font=dict(color=colors['ink-body']),
                annotations=[dict(text="Please select Program and Study first",
                                xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)])
            return empty_fig, html.Div("No data selected"), None, None, ''

        # Filter data based on selections
        # Create two versions: one for the table (all accounts), one for the chart (filtered by account)
        filtered_df_for_table = core_df.copy()
        filtered_df = core_df.copy()

        # Apply program filter to both
        if 'ALL' not in selected_programs:
            filtered_df_for_table = filtered_df_for_table[filtered_df_for_table['P_Code'].isin(selected_programs)]
            filtered_df = filtered_df[filtered_df['P_Code'].isin(selected_programs)]

        # Apply study filter to both
        if 'ALL' not in selected_studies:
            filtered_df_for_table = filtered_df_for_table[filtered_df_for_table['Study_ID'].isin(selected_studies)]
            filtered_df = filtered_df[filtered_df['Study_ID'].isin(selected_studies)]

        # Apply account filter ONLY to chart data (not table data)
        if selected_accounts and 'ALL' not in selected_accounts:
            if 'PF_GROUP' in selected_accounts:
                filtered_df = filtered_df[filtered_df['Account_Clean'].apply(is_program_finance_account)]
            else:
                filtered_df = filtered_df[filtered_df['Account_Clean'].isin(selected_accounts)]

        if len(filtered_df) == 0:
            empty_fig = go.Figure()
            empty_fig.update_layout(
                plot_bgcolor=colors['surface-2'],
                paper_bgcolor=colors['surface-2'],
                font=dict(color=colors['ink-body']),
                annotations=[dict(text="No data found for selection",
                                xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)])
            return empty_fig, html.Div("No data found"), None, None, ''

        # Get historical data (last 8 quarters from actuals)
        hist_quarters = [q for q in quarter_cols if q in filtered_df.columns][-8:]
        hist_data_list = []
        for q in hist_quarters:
            hist_data_list.append({
                'Quarter': q,
                'Amount': filtered_df[q].sum()
            })
        hist_df = pd.DataFrame(hist_data_list)

        # Parse milestone dates
        milestone_dates = {}
        dbl_date = None
        try:
            if fpfv:
                milestone_dates['FPFV'] = pd.to_datetime(fpfv)
            if fpfd:
                milestone_dates['FPFD'] = pd.to_datetime(fpfd)
            if lpfd:
                milestone_dates['LPFD'] = pd.to_datetime(lpfd)
            if lplv:
                milestone_dates['LPLV'] = pd.to_datetime(lplv)
            if dbl:
                dbl_date = pd.to_datetime(dbl)
                milestone_dates['DBL'] = dbl_date
        except:
            pass

        # Calculate forecast duration: Always forecast through DBL + 2 quarters
        from datetime import timedelta, datetime
        import forecast_logic

        if dbl_date and pd.notna(dbl_date):
            # Convert DBL date to its quarter (e.g., Nov 2026 -> FY26 Q4)
            dbl_quarter = forecast_logic.date_to_quarter(dbl_date)

            # Find DBL quarter position in forecast_quarter_cols
            if dbl_quarter in forecast_quarter_cols:
                dbl_index = forecast_quarter_cols.index(dbl_quarter)
                # Forecast through DBL quarter + 2 more quarters
                forecast_periods = min(dbl_index + 3, len(forecast_quarter_cols))  # +3 because we want DBL + 2 additional quarters
                end_quarter = forecast_quarter_cols[forecast_periods - 1] if forecast_periods > 0 else 'N/A'
                logger.info(f"Forecasting to DBL+2Q: DBL is in {dbl_quarter}, forecasting {forecast_periods} quarters through {end_quarter}")
            else:
                # DBL quarter not found in available quarters, calculate from date
                target_end_date = dbl_date + timedelta(days=180)  # ~2 quarters after DBL
                today = datetime.now()
                quarters_to_target = max(4, int((target_end_date - today).days / 90))
                forecast_periods = min(quarters_to_target, len(forecast_quarter_cols))
                logger.info(f"DBL quarter {dbl_quarter} not in forecast range, calculating from date: {forecast_periods} quarters")
        else:
            # No DBL provided: default to 24 quarters (6 years)
            forecast_periods = min(24, len(forecast_quarter_cols))
            logger.info(f"No DBL date provided, defaulting to {forecast_periods} quarters")

        # Get future quarters to forecast - ONLY quarters AFTER the last historical quarter
        last_hist_quarter = hist_quarters[-1] if len(hist_quarters) > 0 else None

        # Filter forecast_quarter_cols to only include quarters after last historical
        if last_hist_quarter and last_hist_quarter in forecast_quarter_cols:
            last_hist_idx = forecast_quarter_cols.index(last_hist_quarter)
            available_future_qtrs = forecast_quarter_cols[last_hist_idx + 1:]  # Start AFTER last historical
        else:
            # If we can't find the last historical quarter, use all forecast quarters
            available_future_qtrs = forecast_quarter_cols

        # Take only the number of periods we need
        future_qtrs = available_future_qtrs[:forecast_periods]

        logger.info(f"Last historical quarter: {last_hist_quarter}, First forecast quarter: {future_qtrs[0] if len(future_qtrs) > 0 else 'N/A'}")

        # Get BSB budget for selected filters
        # Create two versions: one for table (all accounts), one for chart (filtered by account)
        bsb_filtered_for_table = pd.DataFrame()
        bsb_filtered = pd.DataFrame()
        total_budget = 0
        if len(bsb_df) > 0:
            bsb_filtered_for_table = bsb_df.copy()
            bsb_filtered = bsb_df.copy()

            # Apply program/study filters to both
            if 'P_Code' in bsb_filtered.columns and 'ALL' not in selected_programs:
                bsb_filtered_for_table = bsb_filtered_for_table[bsb_filtered_for_table['P_Code'].isin(selected_programs)]
                bsb_filtered = bsb_filtered[bsb_filtered['P_Code'].isin(selected_programs)]
            if 'Study_ID' in bsb_filtered.columns and 'ALL' not in selected_studies:
                bsb_filtered_for_table = bsb_filtered_for_table[bsb_filtered_for_table['Study_ID'].isin(selected_studies)]
                bsb_filtered = bsb_filtered[bsb_filtered['Study_ID'].isin(selected_studies)]

            # Apply account filter ONLY to chart data
            if 'Account_Clean' in bsb_filtered.columns and selected_accounts and 'ALL' not in selected_accounts:
                if 'PF_GROUP' in selected_accounts:
                    bsb_filtered = bsb_filtered[bsb_filtered['Account_Clean'].apply(is_program_finance_account)]
                else:
                    bsb_filtered = bsb_filtered[bsb_filtered['Account_Clean'].isin(selected_accounts)]
            if 'Budget' in bsb_filtered.columns:
                # Apply account-level adjustments to BSB if provided
                if bsb_adjustments and len(bsb_adjustments) > 0:
                    logger.info(f"Applying BSB adjustments: {bsb_adjustments}")
                    # Create adjusted budget column
                    bsb_filtered['Adjusted_Budget'] = bsb_filtered.apply(
                        lambda row: bsb_adjustments.get(row['Account_Clean'], row['Budget'])
                        if pd.notna(row.get('Account_Clean')) else row['Budget'],
                        axis=1
                    )
                    total_budget = bsb_filtered['Adjusted_Budget'].sum()
                    logger.info(f"Total adjusted budget: ${total_budget:,.0f}")
                else:
                    total_budget = bsb_filtered['Budget'].sum()

        # Use user-provided BSB override if available, otherwise use calculated BSB (with adjustments)
        if bsb_total_override and bsb_total_override > 0:
            total_budget = bsb_total_override
            logger.info(f"Using manual BSB override: ${total_budget:,.0f}")
        # If no BSB budget, estimate from historical data (already calculated forecast_periods above)
        elif total_budget == 0:
            total_budget = hist_df['Amount'].sum() * (forecast_periods / 8)
            logger.info(f"No BSB data, estimated from historical: ${total_budget:,.0f}")

        # Calculate actuals spent and open POs
        actuals_spent = hist_df['Amount'].sum()
        po_filtered = pd.DataFrame()
        open_pos = 0
        if len(po_df) > 0:
            po_filtered = po_df.copy()
            if 'P_Code' in po_filtered.columns and 'ALL' not in selected_programs:
                po_filtered = po_filtered[po_filtered['P_Code'].isin(selected_programs)]
            if 'Study_ID' in po_filtered.columns and 'ALL' not in selected_studies:
                po_filtered = po_filtered[po_filtered['Study_ID'].isin(selected_studies)]
            if 'Account_Clean' in po_filtered.columns and selected_accounts and 'ALL' not in selected_accounts:
                if 'PF_GROUP' in selected_accounts:
                    po_filtered = po_filtered[po_filtered['Account_Clean'].apply(is_program_finance_account)]
                else:
                    po_filtered = po_filtered[po_filtered['Account_Clean'].isin(selected_accounts)]
            if 'Open_Amount' in po_filtered.columns:
                open_pos = po_filtered['Open_Amount'].sum()

        # Calculate remaining budget for pattern scaling
        # NOTE: Open POs are PART of the forecast (future spending), not a deduction
        remaining_budget = max(0, total_budget - actuals_spent)
        logger.info(f"Budget breakdown: Total BSB=${total_budget:,.0f}, Actuals=${actuals_spent:,.0f}, Open POs=${open_pos:,.0f} (part of forecast), Remaining to Forecast=${remaining_budget:,.0f}")

        # Generate forecast based on selected method
        show_confidence = 'show' in show_conf if show_conf else False
        show_previous = 'show' in show_prev if show_prev else False
        show_account_lines = 'show' in show_acct_lines if show_acct_lines else False

        # Extract study context for context-aware forecasting
        current_cdu = None
        current_ta = None
        current_study_id = None

        if selected_studies and len(selected_studies) == 1 and 'ALL' not in selected_studies:
            current_study_id = selected_studies[0]
            # Extract study short number to get CDU/TA
            match = re.match(r'\d{4}[-.](\d{3,4})$', str(current_study_id))
            if match:
                study_short = match.group(1)
                matching_row = study_daily_df[study_daily_df['Study_Number_Short'] == study_short]
                if not matching_row.empty:
                    current_cdu = matching_row.iloc[0].get('CDU')
                    current_ta = matching_row.iloc[0].get('Therapeutic_Area')
                    logger.info(f"Study context: {current_study_id}, CDU={current_cdu}, TA={current_ta}")

        # Get normalized spending pattern for ALL methods (background pattern)
        # Extend pattern to include historical quarters for visual comparison
        all_quarters = hist_quarters + future_qtrs
        context_data = {
            'all_actuals_df': core_df,
            'study_daily_df': study_daily_df,
            'current_study_id': current_study_id,
            'current_cdu': current_cdu,
            'current_ta': current_ta
        }
        spending_pattern = forecast_logic.get_normalized_spending_pattern(
            context_data, milestone_dates, all_quarters)

        try:
            # Generate forecast with BSB cap awareness
            if method == 'context':
                # Context-aware forecast using similar studies
                context_data = {
                    'all_actuals_df': core_df,
                    'study_daily_df': study_daily_df,
                    'current_study_id': current_study_id,
                    'current_cdu': current_cdu,
                    'current_ta': current_ta
                }
                generated_forecast_df = forecast_logic.context_aware_forecast(
                    hist_df, future_qtrs, milestone_dates, total_budget, show_confidence,
                    context_data=context_data, actuals_spent=actuals_spent)
            elif method == 'bell':
                generated_forecast_df = forecast_logic.bell_curve_forecast(
                    hist_df, future_qtrs, milestone_dates, total_budget, show_confidence,
                    actuals_spent=actuals_spent)
            elif method == 'linear':
                generated_forecast_df = forecast_logic.linear_forecast(hist_df, future_qtrs, show_confidence)
            elif method == 'polynomial':
                generated_forecast_df = forecast_logic.polynomial_forecast(hist_df, future_qtrs, show_confidence, degree=2)
            elif method == 'exponential':
                generated_forecast_df = forecast_logic.exponential_forecast(hist_df, future_qtrs, show_confidence)
            elif method == 'monte_carlo':
                generated_forecast_df = forecast_logic.monte_carlo_forecast(hist_df, future_qtrs, show_confidence=True)
            elif method == 'tree':
                generated_forecast_df = forecast_logic.tree_forecast(hist_df, future_qtrs, show_confidence)
            else:
                generated_forecast_df = forecast_logic.bell_curve_forecast(
                    hist_df, future_qtrs, milestone_dates, total_budget, show_confidence,
                    actuals_spent=actuals_spent)

            # No need to save - we'll use CORE_Forecast.xlsx data instead

        except Exception as e:
            logger.error(f"Error generating forecast: {e}")
            empty_fig = go.Figure()
            empty_fig.update_layout(
                plot_bgcolor=colors['surface-2'],
                paper_bgcolor=colors['surface-2'],
                font=dict(color=colors['ink-body']),
                annotations=[dict(text=f"Error: {str(e)}",
                                xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)])
            return empty_fig, html.Div(f"Error: {str(e)}"), None

        # Create figure
        fig = go.Figure()

        # ACTUALS LINE
        fig.add_trace(go.Scatter(
            x=hist_df['Quarter'],
            y=hist_df['Amount'],
            mode='lines+markers',
            name='Historical Actuals',
            line=dict(color=colors['brand-cyan'], width=3),
            marker=dict(size=8)))

        # ALWAYS generate account-level forecasts for the snapshot table
        # Use accounts from BOTH actuals AND BSB so all GL accounts are included
        accounts_from_actuals = set(filtered_df_for_table['Account_Clean'].dropna().unique())
        accounts_from_bsb = set(bsb_filtered_for_table['Account_Clean'].dropna().unique()) if len(bsb_filtered_for_table) > 0 and 'Account_Clean' in bsb_filtered_for_table.columns else set()
        # Combine both - all accounts that have either actuals OR BSB budget
        accounts = sorted(list(accounts_from_actuals.union(accounts_from_bsb)))
        logger.info(f"Generating forecast for {len(accounts)} accounts (actuals: {len(accounts_from_actuals)}, BSB: {len(accounts_from_bsb)})")
        plotly_colors = ['#636EFA', '#EF553B', '#00CC96', '#AB63FA', '#FFA15A', '#19D3F3', '#FF6692', '#B6E880']
        account_forecasts = []  # Store all account forecasts for totals calculation

        for idx, account in enumerate(accounts):
            acct_df = filtered_df_for_table[filtered_df_for_table['Account_Clean'] == account]
            acct_hist = pd.DataFrame([{'Quarter': q, 'Amount': acct_df[q].sum()} for q in hist_quarters])

            # Get BSB total limit - use adjustment if available, otherwise use BSB
            # Use unfiltered BSB data so all accounts get their BSB values
            acct_bsb = 0
            if len(bsb_filtered_for_table) > 0 and 'Account_Clean' in bsb_filtered_for_table.columns:
                if bsb_adjustments and account in bsb_adjustments:
                    acct_bsb = bsb_adjustments[account]
                    logger.info(f"Using adjusted BSB for {account}: ${acct_bsb:,.0f}")
                else:
                    acct_bsb = bsb_filtered_for_table[bsb_filtered_for_table['Account_Clean'] == account]['Budget'].sum()
            acct_spent = acct_hist['Amount'].sum()

            # generate forecast per account
            # NOTE: Open POs are not subtracted - they are part of the future forecast
            if method == 'bell':
                acct_forecast = forecast_logic.bell_curve_forecast(acct_hist, future_qtrs, milestone_dates, acct_bsb, False, actuals_spent=acct_spent)
            else:
                acct_forecast = forecast_logic.linear_forecast(acct_hist, future_qtrs, False)

            account_forecasts.append(acct_forecast)

            # Only add account lines to chart if toggle is ON
            if show_account_lines:
                color = plotly_colors[idx % len(plotly_colors)]
                fig.add_trace(go.Scatter(
                    x=acct_forecast['Quarter'],
                    y=acct_forecast['Forecast'],
                    mode='lines+markers',
                    name=f'{account} Forecast (Click to Edit)',
                    line=dict(color=color, width=2, dash='dash'),
                    marker=dict(size=12, line=dict(color='white', width=2), symbol='circle'),
                    hovertemplate='<b>%{x}</b><br>$%{y:,.0f}<br><i>Click to edit</i><extra></extra>',
                    customdata=[{'account': account, 'index': i} for i in range(len(acct_forecast))]))

        # Calculate aggregate forecast from all accounts for totals line
        if len(account_forecasts) > 0:
            aggregate_forecast_df = account_forecasts[0].copy()
            aggregate_forecast_df['Forecast'] = sum(df['Forecast'] for df in account_forecasts)
        else:
            aggregate_forecast_df = generated_forecast_df

        # AGGREGATE PATTERN - Visual roadmap showing expected spending across all phases
        if spending_pattern:
            # Scale pattern to total BSB budget (covers full study lifecycle)
            # This shows the expected spending intensity across historical + forecast periods
            pattern_scaled = [p * total_budget for p in spending_pattern['pattern_y']]
            upper_scaled = [p * total_budget for p in spending_pattern['confidence_upper']]
            lower_scaled = [p * total_budget for p in spending_pattern['confidence_lower']]

            # Golden shaded region showing phase pattern range
            fig.add_trace(go.Scatter(
                x=spending_pattern['pattern_x'],
                y=upper_scaled,
                mode='lines',
                name='Aggregate UPPER',
                line=dict(width=0),
                showlegend=False,
                hoverinfo='skip'))
            fig.add_trace(go.Scatter(
                x=spending_pattern['pattern_x'],
                y=lower_scaled,
                mode='lines',
                name='Phase Pattern Range',
                fill='tonexty',
                fillcolor='rgba(255, 215, 0, 0.25)',  # Golden color with transparency
                line=dict(width=0),
                showlegend=True,
                hovertemplate='<b>%{x}</b><br>Expected Range: $%{y:,.0f} - $%{customdata:,.0f}<extra></extra>',
                customdata=upper_scaled))

        # APOFIS EDITABLE FORECAST LINE
        fig.add_trace(go.Scatter(
            x=generated_forecast_df['Quarter'],
            y=generated_forecast_df['Forecast'],
            mode='lines+markers',
            name='APOFIS Forecast (Click to Edit)',
            line=dict(color='#AB63FA', width=3, dash='dash'),
            marker=dict(
                size=15,  # 15px as requested
                color='#AB63FA',
                line=dict(color='white', width=3),  # thick white outline
                symbol='circle'
            ),
            hovertemplate='<b>%{x}</b><br>$%{y:,.0f}<br><i>Click to edit</i><extra></extra>',
            customdata=list(range(len(generated_forecast_df)))  # store index for editing
        ))

        # display last Forecast (from CORE_Forecast)
        if show_previous:
            try:
                core_forecast_filtered = forecast_df.copy()

                # apply same filters used to generate APOFIS forecast
                if 'ALL' not in selected_programs:
                    core_forecast_filtered = core_forecast_filtered[core_forecast_filtered['P_Code'].isin(selected_programs)]
                if 'ALL' not in selected_studies:
                    core_forecast_filtered = core_forecast_filtered[core_forecast_filtered['Study_ID'].isin(selected_studies)]
                if selected_accounts and 'ALL' not in selected_accounts:
                    if 'PF_GROUP' in selected_accounts:
                        core_forecast_filtered = core_forecast_filtered[core_forecast_filtered['Account_Clean'].apply(is_program_finance_account)]
                    else:
                        core_forecast_filtered = core_forecast_filtered[core_forecast_filtered['Account_Clean'].isin(selected_accounts)]

                # get forecasted qtrs in APOFIS's range to compare
                core_future_data = []
                for quarter in future_qtrs:
                    if quarter in core_forecast_filtered.columns:
                        core_future_data.append({
                            'Quarter': quarter,
                            'Amount': core_forecast_filtered[quarter].sum()})

                # LAST FORECAST LINE
                if len(core_future_data) > 0:
                    core_df_plot = pd.DataFrame(core_future_data)
                    fig.add_trace(go.Scatter(
                        x=core_df_plot['Quarter'],
                        y=core_df_plot['Amount'],
                        mode='lines+markers',
                        name='CORE Forecast',
                        line=dict(color=colors['ink-soft'], width=2, dash='dot'),
                        marker=dict(size=6),
                        opacity=0.7
                    ))
            except Exception as e:
                logger.warning(f"Could not display CORE forecast: {e}")

        # add confidence interval if enabled
        if show_confidence:
            fig.add_trace(go.Scatter(
                x=generated_forecast_df['Quarter'],
                y=generated_forecast_df['Upper_Bound'],
                mode='lines',
                name='Upper Bound',
                line=dict(width=0),
                showlegend=False,
                hoverinfo='skip'))
            fig.add_trace(go.Scatter(
                x=generated_forecast_df['Quarter'],
                y=generated_forecast_df['Lower_Bound'],
                mode='lines',
                name='Confidence Interval',
                fill='tonexty',
                fillcolor='rgba(0, 229, 255, 0.2)',
                line=dict(width=0),
                showlegend=True
            ))

        # running total line (cumulative sum)
        combined_data = []
        # actual quarters
        for _, row in hist_df.iterrows():
            combined_data.append({'Quarter': row['Quarter'], 'Amount': row['Amount']})
        # forecast quarters - use aggregate from account forecasts (always generated now)
        if len(account_forecasts) > 0:
            for _, row in aggregate_forecast_df.iterrows():
                combined_data.append({'Quarter': row['Quarter'], 'Amount': row['Forecast']})
        elif 'Forecast' in generated_forecast_df.columns:
            for _, row in generated_forecast_df.iterrows():
                combined_data.append({'Quarter': row['Quarter'], 'Amount': row['Forecast']})

        combined_df = pd.DataFrame(combined_data)
        combined_df['Running_Total'] = combined_df['Amount'].cumsum()

        fig.add_trace(go.Scatter(
            x=combined_df['Quarter'],
            y=combined_df['Running_Total'],
            mode='lines',
            name='Running Total',
            line=dict(color=colors['brand-green'], width=3),
            fill='tozeroy',
            fillcolor='rgba(0, 255, 156, 0.08)',
            yaxis='y2'))

        # Add milestone markers (vertical red dashed lines)
        milestone_labels = {
            'FPFV': fpfv,
            'FPFD': fpfd,
            'LPFD': lpfd,
            'LPLV': lplv,
            'DBL': dbl}

        # Collect all quarters in the plot for x-axis positioning
        all_quarters = combined_df['Quarter'].tolist()

        milestone_shapes = []
        milestone_annotations = []

        for label, date_val in milestone_labels.items():
            if date_val:
                try:
                    # Convert date string to quarter format (e.g., "2024Q1")
                    from datetime import datetime
                    if isinstance(date_val, str):
                        # Handle different date string formats
                        try:
                            milestone_date = datetime.fromisoformat(date_val)
                        except:
                            milestone_date = datetime.strptime(date_val, '%Y-%m-%d')
                    else:
                        milestone_date = date_val

                    quarter_num = (milestone_date.month - 1) // 3 + 1
                    milestone_quarter = f"{milestone_date.year}Q{quarter_num}"

                    # Find position in quarter list
                    if milestone_quarter in all_quarters:
                        x_pos = all_quarters.index(milestone_quarter)

                        # Add vertical line using shapes
                        milestone_shapes.append(dict(
                            type='line',
                            x0=x_pos,
                            x1=x_pos,
                            y0=0,
                            y1=1,
                            yref='paper',
                            line=dict(color='red', width=2, dash='dash')
                        ))

                        # Add annotation
                        milestone_annotations.append(dict(
                            x=x_pos,
                            y=1,
                            yref='paper',
                            text=label,
                            showarrow=False,
                            font=dict(size=10, color='red'),
                            yshift=10
                        ))
                except Exception as e:
                    logger.warning(f"Could not add milestone marker for {label}: {e}")

        fig.update_layout(
            plot_bgcolor=colors['surface-2'],
            paper_bgcolor=colors['surface-2'],
            font=dict(color=colors['ink-body']),
            xaxis=dict(showgrid=True, gridcolor=colors['line-soft'], title='Quarter'),
            yaxis=dict(
                showgrid=True,
                gridcolor=colors['line-soft'],
                title='Quarterly Amount ($)',
                tickformat='$,.0f'
            ),
            yaxis2=dict(
                title=dict(text='Running Total ($)', font=dict(color=colors['brand-green'])),
                tickfont=dict(color=colors['brand-green']),
                tickformat='$,.0f',
                overlaying='y',
                side='right'
            ),
            margin=dict(l=60, r=80, t=40, b=80),
            showlegend=True,
            legend=dict(
                orientation='h',
                yanchor='bottom',
                y=1.02,
                xanchor='right',
                x=1,
                font=dict(color=colors['ink-body'])
            ),
            hovermode='closest',  # Changed to 'closest' for better point clicking
            dragmode='pan',  # Enable pan/zoom
            shapes=milestone_shapes,
            annotations=milestone_annotations
        )

        # Create monthly tabular view for Excel-friendly export (ALWAYS with account breakdown)
        # Convert quarterly forecast to monthly breakdown
        monthly_df = forecast_logic.quarterly_to_monthly(generated_forecast_df)

        if len(account_forecasts) > 0:
            # ACCOUNT BREAKOUT: Create table with accounts as rows, months as columns
            account_monthly_data = []

            for idx, account in enumerate(accounts):
                acct_df = filtered_df[filtered_df['Account_Clean'] == account]
                acct_hist = pd.DataFrame([{'Quarter': q, 'Amount': acct_df[q].sum()} for q in hist_quarters])

                # Get account-specific forecast
                acct_forecast = account_forecasts[idx]

                # Convert to monthly
                acct_monthly = forecast_logic.quarterly_to_monthly(acct_forecast)

                # Calculate actualized (historical) spend for this account
                acct_actualized = acct_hist['Amount'].sum()

                # Build row: Account name + monthly values
                row_data = {'Account': account}
                for _, month_row in acct_monthly.iterrows():
                    row_data[month_row['Month']] = month_row['Forecast']

                # Store actualized for later addition
                row_data['_actualized'] = acct_actualized
                account_monthly_data.append(row_data)

            # Create DataFrame
            if len(account_monthly_data) > 0:
                monthly_table_df = pd.DataFrame(account_monthly_data)

                # Get month columns (exclude Account and temporary _actualized)
                month_columns = [col for col in monthly_table_df.columns if col not in ['Account', '_actualized']]

                # Add FORECAST TOTAL column (sum across all forecast months for each account)
                monthly_table_df['Forecast Total'] = monthly_table_df[month_columns].sum(axis=1)

                # Add ACTUALIZED column (rename from temporary)
                monthly_table_df['Actualized'] = monthly_table_df['_actualized']
                monthly_table_df = monthly_table_df.drop('_actualized', axis=1)

                # Add TOTAL column (Forecast Total + Actualized)
                monthly_table_df['Total'] = monthly_table_df['Forecast Total'] + monthly_table_df['Actualized']

                # Add TOTAL row (sum across all accounts)
                total_row = {'Account': 'TOTAL'}
                for month_col in month_columns:
                    total_row[month_col] = monthly_table_df[month_col].sum()
                total_row['Forecast Total'] = monthly_table_df['Forecast Total'].sum()
                total_row['Actualized'] = monthly_table_df['Actualized'].sum()
                total_row['Total'] = monthly_table_df['Total'].sum()

                monthly_table_df = pd.concat([monthly_table_df, pd.DataFrame([total_row])], ignore_index=True)

                # All columns in order: Account, months, Forecast Total, Actualized, Total
                all_value_columns = month_columns + ['Forecast Total', 'Actualized', 'Total']

                table = dash_table.DataTable(
                    data=monthly_table_df.to_dict('records'),
                    columns=[{'name': 'Account', 'id': 'Account'}] +
                            [{'name': col, 'id': col, 'type': 'numeric', 'format': {'specifier': '$,.0f'}}
                             for col in all_value_columns],
                    style_table={
                        'backgroundColor': colors['surface-2'],
                        'overflowX': 'auto',
                        'width': '100%'
                    },
                    style_cell={
                        'backgroundColor': colors['surface-2'],
                        'color': colors['ink-body'],
                        'border': f"1px solid {colors['line-soft']}",
                        'textAlign': 'right',
                        'padding': '8px'
                    },
                    style_cell_conditional=[
                        {'if': {'column_id': 'Account'},
                         'textAlign': 'left',
                         'fontWeight': 'bold',
                         'minWidth': '200px',
                         'position': 'sticky',
                         'left': 0,
                         'backgroundColor': colors['surface-2']},
                        {'if': {'column_id': 'Actualized'},
                         'fontWeight': 'bold',
                         'backgroundColor': colors['surface-0']},
                        {'if': {'column_id': 'Forecast Total'},
                         'fontWeight': 'bold',
                         'backgroundColor': colors['surface-0']},
                        {'if': {'column_id': 'Total'},
                         'fontWeight': 'bold',
                         'backgroundColor': colors['brand-cyan'],
                         'color': 'white'}
                    ],
                    style_header={
                        'backgroundColor': COLORS['surface-0'],
                        'color': colors['ink-strong'],
                        'fontWeight': 'bold',
                        'border': f"1px solid {colors['line-soft']}",
                        'textAlign': 'center'
                    },
                    style_data_conditional=[
                        {'if': {'column_id': 'Account'},
                         'fontWeight': 'bold'},
                        {'if': {'filter_query': '{Account} = "TOTAL"'},
                         'fontWeight': 'bold',
                         'backgroundColor': colors['surface-0'],
                         'color': colors['brand-cyan']}
                    ],
                    fixed_columns={'headers': True, 'data': 1}
                )
            else:
                table = html.Div("No account data available")
        else:
            # Fallback if no accounts found
            table = html.Div("No forecast data available")

        # Store the forecast DataFrame as JSON for editing
        forecast_data = {
            'quarters': generated_forecast_df['Quarter'].tolist(),
            'values': generated_forecast_df['Forecast'].tolist()
        }

        # Store monthly table data for Excel export (with Actualized and totals)
        monthly_export_data = None
        if len(account_forecasts) > 0:
            # Store account-level monthly breakdown with actualized and totals
            export_rows = []
            month_totals = {}  # Track totals for each month
            total_actualized = 0  # Track total actualized across all accounts

            for idx, account in enumerate(accounts):
                acct_df = filtered_df[filtered_df['Account_Clean'] == account]
                acct_hist = pd.DataFrame([{'Quarter': q, 'Amount': acct_df[q].sum()} for q in hist_quarters])
                acct_actualized = acct_hist['Amount'].sum()
                total_actualized += acct_actualized

                acct_forecast = account_forecasts[idx]
                acct_monthly = forecast_logic.quarterly_to_monthly(acct_forecast)

                row_data = {'Account': account}
                forecast_total = 0
                for _, month_row in acct_monthly.iterrows():
                    month_name = month_row['Month']
                    month_value = month_row['Forecast']
                    row_data[month_name] = month_value
                    forecast_total += month_value

                    # Accumulate month totals
                    if month_name not in month_totals:
                        month_totals[month_name] = 0
                    month_totals[month_name] += month_value

                row_data['Forecast Total'] = forecast_total
                row_data['Actualized'] = acct_actualized
                row_data['Total'] = forecast_total + acct_actualized
                export_rows.append(row_data)

            # Add TOTAL row
            total_row = {'Account': 'TOTAL'}
            grand_forecast_total = 0
            for month_name, month_total in month_totals.items():
                total_row[month_name] = month_total
                grand_forecast_total += month_total
            total_row['Forecast Total'] = grand_forecast_total
            total_row['Actualized'] = total_actualized
            total_row['Total'] = grand_forecast_total + total_actualized

            export_rows.append(total_row)
            monthly_export_data = export_rows
        else:
            # Fallback if no accounts
            if len(monthly_df) > 0:
                row_data = {'Account': 'Total Forecast'}
                row_total = 0
                for _, month_row in monthly_df.iterrows():
                    month_value = month_row['Forecast']
                    row_data[month_row['Month']] = month_value
                    row_total += month_value
                row_data['Total'] = row_total
                monthly_export_data = [row_data]

        return fig, table, forecast_data, monthly_export_data, ''

    @app.callback(
        [Output('proxy-forecast-chart', 'figure'),
         Output('proxy-similar-studies-table', 'children')],
        Input('generate-proxy-btn', 'n_clicks'),
        [State('proxy-indication', 'value'),
         State('proxy-phase', 'value'),
         State('proxy-duration', 'value'),
         State('proxy-budget', 'value'),
         State('proxy-enrollment', 'value'),
         State('theme-store', 'data')],
        prevent_initial_call=True
    )
    def generate_proxy_forecast(n_clicks, indication, phase, duration, budget, enrollment, theme):
        """Generate proxy forecast based on indication, phase, timeline, and budget."""
        import forecast_logic
        from dash import dash_table

        # Get colors for current theme
        colors = get_colors(theme)

        if not n_clicks or not budget:
            empty_fig = go.Figure()
            empty_fig.update_layout(
                plot_bgcolor=colors['surface-2'],
                paper_bgcolor=colors['surface-2'],
                font=dict(color=colors['ink-body']),
                annotations=[dict(text="Enter parameters and click Generate",
                                xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)])
            return empty_fig, html.Div("Enter parameters above")

        try:
            # Convert months to quarters
            num_quarters = int(duration / 3)

            # Generate quarterly labels
            quarters = [f"Q{i+1}" for i in range(num_quarters)]

            # Use bell curve distribution to spread budget over timeline
            # Create synthetic milestone dates based on typical study timeline
            from datetime import datetime, timedelta
            today = datetime.now()

            # Typical timeline: 20% ramp-up, 50% enrollment, 20% follow-up, 10% closeout
            rampup_months = int(duration * 0.2)
            enrollment_months = int(duration * 0.5)
            followup_months = int(duration * 0.2)

            milestone_dates = {
                'FPFV': today,
                'FPFD': today + timedelta(days=30 * rampup_months),
                'LPFD': today + timedelta(days=30 * (rampup_months + enrollment_months)),
                'LPLV': today + timedelta(days=30 * (rampup_months + enrollment_months + followup_months)),
                'DBL': today + timedelta(days=30 * duration)
            }

            # Create dummy historical data (not used for proxy, but required by function)
            hist_df = pd.DataFrame({'Quarter': ['Q1', 'Q2'], 'Amount': [0, 0]})

            # Generate forecast
            forecast_df = forecast_logic.bell_curve_forecast(
                hist_df, quarters, milestone_dates, budget, show_confidence=True)

            # Create figure
            fig = go.Figure()

            # Add forecast line
            fig.add_trace(go.Scatter(
                x=forecast_df['Quarter'],
                y=forecast_df['Forecast'],
                mode='lines+markers',
                name='Median Forecast',
                line=dict(color=colors['brand-cyan'], width=3),
                marker=dict(size=8)
            ))

            # Add confidence interval
            fig.add_trace(go.Scatter(
                x=forecast_df['Quarter'],
                y=forecast_df['Upper_Bound'],
                mode='lines',
                name='Upper Bound',
                line=dict(width=0),
                showlegend=False,
                hoverinfo='skip'
            ))
            fig.add_trace(go.Scatter(
                x=forecast_df['Quarter'],
                y=forecast_df['Lower_Bound'],
                mode='lines',
                name='25th-75th Percentile',
                fill='tonexty',
                fillcolor='rgba(0, 229, 255, 0.2)',
                line=dict(width=0),
                showlegend=True
            ))

            fig.update_layout(
                plot_bgcolor=colors['surface-2'],
                paper_bgcolor=colors['surface-2'],
                font=dict(color=colors['ink-body']),
                xaxis=dict(showgrid=True, gridcolor=colors['line-soft'], title='Quarter'),
                yaxis=dict(showgrid=True, gridcolor=colors['line-soft'], title='Quarterly Spend ($)', tickformat='$,.0f'),
                margin=dict(l=60, r=20, t=20, b=80),
                showlegend=True,
                legend=dict(
                    orientation='h',
                    yanchor='bottom',
                    y=1.02,
                    xanchor='right',
                    x=1,
                    font=dict(color=colors['ink-body'])
                ),
                hovermode='x unified'
            )

            # Find similar studies from study_daily_df
            similar_studies = []
            if indication and phase:
                matches = study_daily_df[
                    (study_daily_df['Indication'].str.contains(indication, case=False, na=False)) &
                    (study_daily_df['Phase'].str.contains(phase, case=False, na=False))
                ]

                if len(matches) > 0:
                    for idx, row in matches.head(5).iterrows():
                        similar_studies.append({
                            'Study': row.get('Study Number', 'Unknown'),
                            'Phase': row.get('Phase', '—'),
                            'Indication': row.get('Indication', '—'),
                            'Enrollment': int(row.get('# Enrollment (Planned)', 0)) if pd.notna(row.get('# Enrollment (Planned)')) else '—'
                        })

            if len(similar_studies) > 0:
                table = dash_table.DataTable(
                    data=similar_studies,
                    columns=[{'name': c, 'id': c} for c in ['Study', 'Phase', 'Indication', 'Enrollment']],
                    style_table={'backgroundColor': colors['surface-2']},
                    style_cell={
                        'backgroundColor': colors['surface-2'],
                        'color': colors['ink-body'],
                        'border': f"1px solid {colors['line-soft']}",
                        'textAlign': 'left',
                        'padding': '12px'
                    },
                    style_header={
                        'backgroundColor': COLORS['surface-0'],
                        'color': colors['ink-strong'],
                        'fontWeight': 'bold',
                        'border': f"1px solid {colors['line-soft']}"
                    }
                )
            else:
                table = html.Div("No similar studies found in database",
                               style={'color': colors['ink-soft'], 'padding': '20px'})

            return fig, table

        except Exception as e:
            logger.error(f"Error generating proxy forecast: {e}")
            empty_fig = go.Figure()
            empty_fig.update_layout(
                plot_bgcolor=colors['surface-2'],
                paper_bgcolor=colors['surface-2'],
                font=dict(color=colors['ink-body']),
                annotations=[dict(text=f"Error: {str(e)}",
                                xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)])
            return empty_fig, html.Div(f"Error: {str(e)}"), None

    # POINT EDITOR - Show editor when clicking forecast point ===
    @app.callback(
        [Output('point-editor', 'children'),
         Output('editing-point-store', 'data')],
        [Input('forecast-chart', 'clickData'),
         Input({'type': 'update-point-btn', 'index': ALL}, 'n_clicks'),
         Input({'type': 'cancel-edit-btn', 'index': ALL}, 'n_clicks')],
        [State('forecast-chart', 'figure'),
         State('theme-store', 'data')],
        prevent_initial_call=True
    )
    def show_point_editor(click_data, update_clicks, cancel_clicks, current_figure, theme):
        """Show editor panel when user clicks a forecast point."""
        from dash import ctx

        # Get colors for current theme
        colors = get_colors(theme)

        # Check what triggered this callback
        if not ctx.triggered_id:
            return no_update, no_update

        triggered_id = ctx.triggered_id

        # If Update or Cancel button clicked, hide the editor
        if isinstance(triggered_id, dict):
            if triggered_id.get('type') in ['update-point-btn', 'cancel-edit-btn']:
                return html.Div(), None

        # Otherwise, it was a click on the chart - show the editor
        if not click_data:
            return html.Div(), None

        # Get clicked point info
        point = click_data['points'][0]
        curve_number = point.get('curveNumber')
        point_number = point.get('pointNumber')
        x_val = point.get('x')
        y_val = point.get('y')

        # Check if it's any forecast trace (editable)
        if current_figure and 'data' in current_figure:
            clicked_trace = current_figure['data'][curve_number]
            trace_name = clicked_trace.get('name', '')

            # Check if it's an editable trace (any forecast line with "Click to Edit")
            if 'Forecast' in trace_name and 'Click to Edit' in trace_name:
                # Extract account name if it's an account-specific forecast
                account_name = None
                if ' Forecast (Click to Edit)' in trace_name:
                    account_name = trace_name.replace(' Forecast (Click to Edit)', '')

                # Show editor panel
                title_text = f'Editing {account_name} - {x_val}' if account_name else f'Editing {x_val}'
                editor = html.Div([
                    dbc.Card([
                        dbc.CardBody([
                            html.Div([
                                html.Strong(title_text, style={'color': colors['brand-cyan']}),
                                html.Span(f' - Current: ${y_val:,.0f}', style={'color': colors['ink-body'], 'marginLeft': '10px'})
                            ]),
                            html.Div([
                                dcc.Input(
                                    id={'type': 'point-value-input', 'index': f'{curve_number}_{point_number}'},
                                    type='number',
                                    value=y_val,
                                    style={'width': '200px', 'marginRight': '10px', 'marginTop': '10px'}
                                ),
                                dbc.Button('Update', id={'type': 'update-point-btn', 'index': f'{curve_number}_{point_number}'},
                                          color='success', size='sm', style={'marginTop': '10px', 'marginRight': '5px'}),
                                dbc.Button('Cancel', id={'type': 'cancel-edit-btn', 'index': f'{curve_number}_{point_number}'},
                                          color='secondary', size='sm', outline=True, style={'marginTop': '10px'})
                            ], style={'marginTop': '10px'})
                        ])
                    ], style={'backgroundColor': COLORS['surface-0'], 'border': f'2px solid {COLORS["brand-cyan"]}'})
                ], style={'marginBottom': '10px'})

                # Store which point is being edited
                editing_info = {
                    'curve': curve_number,
                    'point': point_number,
                    'quarter': x_val,
                    'original_value': y_val,
                    'account': account_name
                }
                return editor, editing_info

        return html.Div(), None

    # UPDATE THE FORECAST DATA AND REGENERATE CHART
    @app.callback(
        [Output('custom-forecast-store', 'data', allow_duplicate=True),
         Output('forecast-chart', 'figure', allow_duplicate=True)],
        Input({'type': 'update-point-btn', 'index': ALL}, 'n_clicks'),
        [State({'type': 'point-value-input', 'index': ALL}, 'value'),
         State('editing-point-store', 'data'),
         State('custom-forecast-store', 'data'),
         State('forecast-chart', 'figure'),
         State('theme-store', 'data')],
        prevent_initial_call=True
    )
    def update_forecast_data(update_clicks, input_values, store_data, forecast_data, current_figure, theme):
        """Update the forecast data and regenerate the chart."""
        from dash import ctx

        # Get colors for current theme
        colors = get_colors(theme)

        if not ctx.triggered_id or not isinstance(ctx.triggered_id, dict) or ctx.triggered_id.get('type') != 'update-point-btn':
            return no_update, no_update

        if not input_values or not input_values[0] or not store_data or not forecast_data:
            return no_update, no_update

        point_idx = store_data['point']
        new_value = float(input_values[0])

        logger.info(f"UPDATE: Changing point {point_idx} to ${new_value:,.0f}")

        # Update the stored forecast values
        forecast_data['values'][point_idx] = new_value

        # Rebuild the ENTIRE figure from scratch with updated data
        new_fig = go.Figure()

        # Copy all traces from current figure
        for trace_data in current_figure['data']:
            # If it's the APOFIS forecast trace, use updated values
            if 'APOFIS Forecast' in trace_data.get('name', ''):
                new_trace = go.Scatter(
                    x=forecast_data['quarters'],
                    y=forecast_data['values'],
                    mode=trace_data.get('mode'),
                    name=trace_data.get('name'),
                    line=trace_data.get('line'),
                    marker=trace_data.get('marker'),
                    hovertemplate=trace_data.get('hovertemplate'),
                    customdata=trace_data.get('customdata')
                )
                new_fig.add_trace(new_trace)
                logger.info(f"Rebuilt APOFIS trace with updated values")
            else:
                # Copy other traces as-is
                new_fig.add_trace(trace_data)

        # Copy layout
        if 'layout' in current_figure:
            new_fig.update_layout(current_figure['layout'])

        return forecast_data, new_fig

    # EXPORT FORECAST TO EXCEL ===================================
    @app.callback(
        Output('download-forecast', 'data'),
        Input('export-forecast-btn', 'n_clicks'),
        State('monthly-table-store', 'data'),
        prevent_initial_call=True
    )
    def export_forecast_to_excel(n_clicks, monthly_data):
        """Export the monthly tabular forecast to Excel."""
        if not n_clicks or not monthly_data:
            return no_update

        try:
            # Convert monthly data to DataFrame
            df_export = pd.DataFrame(monthly_data)

            if df_export.empty:
                logger.warning("No forecast data to export")
                return no_update

            # Create Excel file in memory using openpyxl
            import io
            from openpyxl.styles import numbers, Alignment, Font, PatternFill
            output = io.BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export.to_excel(writer, sheet_name='Monthly Forecast', index=False)

                # Get workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Monthly Forecast']

                # Find column indices for special columns
                header_row = [cell.value for cell in worksheet[1]]
                actualized_col = header_row.index('Actualized') + 1 if 'Actualized' in header_row else None
                forecast_total_col = header_row.index('Forecast Total') + 1 if 'Forecast Total' in header_row else None
                total_col = header_row.index('Total') + 1 if 'Total' in header_row else None

                # Color fills
                light_gray_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
                cyan_fill = PatternFill(start_color='00E5FF', end_color='00E5FF', fill_type='solid')

                # Format the sheet
                # 1. Bold header row
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # 2. Format Account column (A) - left-aligned, bold
                worksheet.column_dimensions['A'].width = 25
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=1)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left')

                # 3. Format all numerical columns (B onwards) - plain number format, right-aligned
                for col in range(2, worksheet.max_column + 1):
                    # Auto-adjust column width
                    col_letter = worksheet.cell(row=1, column=col).column_letter
                    worksheet.column_dimensions[col_letter].width = 15

                    # Format cells
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.number_format = '0'  # Plain number format (no $ or commas)
                        cell.alignment = Alignment(horizontal='right')

                        # Highlight special columns
                        if col == actualized_col or col == forecast_total_col:
                            cell.fill = light_gray_fill
                            cell.font = Font(bold=True)
                        elif col == total_col:
                            cell.fill = cyan_fill
                            cell.font = Font(bold=True)

                # 4. Bold and highlight TOTAL row
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=worksheet.max_row, column=col)
                    cell.font = Font(bold=True, color='00E5FF')

                # 5. Freeze panes (freeze first row and first column)
                worksheet.freeze_panes = 'B2'

            output.seek(0)

            # Generate filename with timestamp
            from datetime import datetime
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'APOFIS_Forecast_{timestamp}.xlsx'

            logger.info(f"Exporting monthly forecast with {len(df_export)} accounts and {len(df_export.columns)-1} months")

            return dcc.send_bytes(output.getvalue(), filename)

        except Exception as e:
            import traceback
            logger.error(f"Error exporting forecast: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return no_update

    # SHOW INSTRUCTIONS FOR ALL FORECASTS =======================
    @app.callback(
        Output('forecast-instructions', 'children'),
        [Input('forecast-method', 'value'),
         Input('theme-store', 'data')]
    )
    def show_forecast_instructions(method, theme):
        """Display instructions for interactive forecasting."""
        colors = get_colors(theme)
        return html.Div([
            html.I(className='fas fa-hand-pointer', style={'marginRight': '8px', 'color': colors['brand-cyan']}),
            html.Span('Click any forecast point to edit its value. ',
                     style={'color': colors['ink-body'], 'fontWeight': 'bold'}),
            html.Span('Golden shaded area shows expected spending pattern based on milestone phases - compare your actuals to see if you\'re tracking typical patterns.',
                     style={'color': colors['ink-soft'], 'fontStyle': 'italic'})
        ], style={'padding': '10px', 'backgroundColor': colors['surface-0'], 'borderRadius': '4px',
                 'border': f'1px solid {colors["line-soft"]}'})

    logger.info("All callbacks registered successfully")

    logger.info('BSB callbacks registered')
